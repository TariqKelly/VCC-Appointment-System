import openpyxl as pxl
from openpyxl import load_workbook
import pandas as pd
from pandas.core.frame import DataFrame
from pandas.core.series import Series
import numpy as np




def create_database(path):
    
    workbook = pxl.Workbook()
    patients = workbook.active
    patients.title = "Contact Details"
    patients["A1"] = "ID #"
    patients["B1"] = "First Name"
    patients["C1"] = "Last Name"
    patients["D1"] = "Tel. Number"
    patients["E1"] = "Email Address"
    workbook.save(path)

def add_patient():

    ID = input("Enter Patient ID: ")

    first_name = input("Enter Patient's First Name: ")
    while not first_name.isalpha():
        print("Please enter only alphabets for the first name")
        first_name = input("Enter Patient's First Name: ")
    
    
    last_name = input("Enter Patient's Last Name: ")
    while not last_name.isalpha():
        print("Please enter only alphabets for the last name")
        first_name =input("Enter Patient's Last Name: ")

    tel_num = input("Enter Patient's telephone number: ")
    while not tel_num.isdigit():
        print("Invalid telephone number. Should only contain numbers.\n Try again: ")
        tel_num = input()
    while len(tel_num) < 10:
        dif = 10 - len(tel_num)
        print(f"Telephone number is missing {dif} digits.\n")
        tel_num = input("Re-Enter Number: ")
    while len(tel_num) > 10:
        ext = len(tel_num) - 10
        print(f"Telephone number has {ext} extra digits.\n")
        tel_num = input("Re-Enter Number: ")

    else:
        tlnm_1 = tel_num[:3]
        tlnm_2 = tel_num[3:6]
        tlnm_3 = tel_num[6:]
        tel_num = tlnm_1 + "-" + tlnm_2 + "-" + tlnm_3

    email_add = input("Enter Patient's Email Address: ")
    val_em = ("@", ".com")
    if email_add == "":
        email_add = "Not Applicable"
    else:
        for i in val_em:
            while i not in email_add:
                email_add = input(f"Email address missing {i}. Please enter again: ")

    new_patient = [[ID, first_name, last_name, tel_num, email_add]]

    for info in new_patient:
        page.append(info)
    
    wb.save(filename=workbook)
    drop_duplicates()

def find_patient():

    df = pd.read_excel(workbook)
    result = []
    choice = input("Enter string to search by: ")
    df.dropna(inplace=True)
    for column in df:
        result.append(column)
    for pos in result[1:3]:
        df2 = df[df[pos].str.contains(choice)]
        if df2.empty:
            print(f"Not found in {pos}")
            pass
        else: 
            print(f"Found in {pos}\n", df2)
            pass
        
    """search = input("Enter ID to search: ")
    for data in sheet["A"]:
        if data.value == search:
            for row in page.iter_rows(min_row=data.row, min_col=1, max_row=data.row, max_col=5):
                for cell in row:
                    result.append(cell.value)
            print(result[1:])
    return result"""
 
def findpos(id_num):
    idx = 0
    for data in sheet['A']:
        if data.value == str(id_num):
            idx = data.row
    return (idx)
             

def update_patient():

    print("Select the patient record to update\n")
    patient_info = find_patient()
    print(patient_info)
    page = wb.active
    sheet = wb["Contact Details"]
                
    print("[Select 1 to change Patient's First Name]\n \
            [Select 2 to change Patient's Last Name]\n \
             [Select 3 to change Patient's Telephone Number]\n \
              [Select 4 to change Patient's Email Address]\n")
    
    select = input("Choice: ")
    
    if   select == "1":
        patient_info[1] = input("Enter New First Name: ")

        for data in sheet["A"]:
            for row in page.iter_rows(min_row=findpos(patient_info[0]), min_col=2, max_row=findpos(patient_info[0]), max_col=2):
                for cell in row:
                    cell.value = patient_info[1]
                    wb.save(filename=workbook)

    elif select == "2":
        for data in sheet["A"]:
            patient_info[2] = input("Enter New Last Name: ")
            for row in page.iter_rows(min_row=findpos(patient_info[0]), min_col=3, max_row=findpos(patient_info[0]), max_col=3):
                for cell in row:
                    cell.value = patient_info[2]
                    wb.save(filename=workbook)

    elif select == "3":
        patient_info[3] = input("Enter New Telephone Number: ")
        for data in sheet["A"]:
            for row in page.iter_rows(min_row=findpos(patient_info[0]), min_col=4, max_row=findpos(patient_info[0]), max_col=4):
                for cell in row:
                    cell.value = patient_info[3]
                    wb.save(filename=workbook)
                    
    elif select == "4":
        patient_info[4] = input("Enter New Email Address: ")
        for data in sheet["A"]:
            for row in page.iter_rows(min_row=findpos(patient_info[0]), min_col=5, max_row=findpos(patient_info[0]), max_col=5):
                for cell in row:
                    cell.value = patient_info[4]
                    wb.save(filename=workbook)

def delete_patient():
    choice = input("Choose patient ID to be deleted: ")
    page.delete_rows(findpos(choice))
    wb.save(filename=workbook)

def sort():
    sort_by = input("Order to sort by: ")
    df = pd.read_excel(workbook, sheet_name= "Contact Details")
    print("Original Database")
    print(df)

    df = df.sort_values(sort_by)

    writer = pd.ExcelWriter(workbook_sorted)
    df.to_excel(writer, index = False)
    writer.save()
    print ("Sorted")
    print(df)

    wb.save(filename=workbook)
    #wb.save(filename = workbook_sorted)

def drop_duplicates():
    
    new_add = pd.read_excel(workbook, sheet_name="Contact Details")
    dupe = new_add.duplicated(subset=["First Name", "Last Name", "Tel. Number", "Email Address"], keep = "first")
    ls = list(dupe)
    for bool in ls:
        if bool == True:
            print("Patient already in databse\n")
    drop=new_add.drop_duplicates(subset=["First Name", "Last Name", "Tel. Number", "Email Address"], keep = "first")

    writer = pd.ExcelWriter(workbook)
    drop.to_excel(writer, sheet_name="Contact Details", index = False)
    writer.save()
    

workbook = "Patients_Database.xlsx"
workbook_sorted = "Patients_Database_Sorted.xlsx"
wb = load_workbook(workbook)
sheet = wb["Contact Details"]
page = wb.active

add_patient()
#find_patient()
                    
"""if __name__ == "__main__":
    #create_database("Patients_Database_Sorted.xlsx")
    #for i in range (0, 4):
    #add_patient()
    #find_patient()
    #update_patient()
    #delete_patient()
    #sort()
    #drop_duplicates()
    pass"""
