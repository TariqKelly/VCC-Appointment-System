import openpyxl as pxl
from openpyxl import load_workbook
from openpyxl import workbook
import pandas as pd

workbook = "Patients_Database.xlsx"
wb = load_workbook(workbook)
sheet = wb["Contact Details"]
page = wb.active

def create_database(path):
    
    workbook = pxl.Workbook()
    patients = workbook.active
    patients.title = "Contact Details"
    patients["A1"] = "ID #"
    patients["B1"] = "First Name"
    patients["C1"] = "Last Name"
    patients["D1"] = "Tel. Number"
    patients["E1"] = "Email Address"
    workbook.save(path)

def add_patient():
    
    ID = str(input("Enter Patient ID: "))
    first_name = str(input("Enter Patient's First Name: "))
    last_name = str(input("Enter Patient's Last Name: "))
    tel_number = str(input("Enter Patient's telephone number: "))
    email_add = str(input("Enter Patient's Email Address: "))

    new_patient = [[ID, first_name, last_name, tel_number, email_add]]

    for info in new_patient:
        page.append(info)

    wb.save(filename=workbook)

def find_patient():

    workbook = "Patients_Database.xlsx"
    wb = load_workbook(workbook)
    sheet = wb["Contact Details"]
    page = wb.active
    result = []
    search = str(input("Enter ID to search: "))
    for data in sheet['A']:
        if data.value == search:
            for row in page.iter_rows(min_row=data.row, min_col=1, max_row=data.row, max_col=5):
                for cell in row:
                    result.append(cell.value)
            print(result[1:])
    return result
 
def findpos(id_num):
    idx = 0
    for data in sheet['A']:
        if data.value == str(id_num):
            idx = data.row
    return (idx)
             

def update_patient():

    print("Select the patient record to update\n")
    patient_info = find_patient()
    print(patient_info)
    page = wb.active
    sheet = wb["Contact Details"]
                
    print("[Select 1 to change Patient's First Name]\n \
            [Select 2 to change Patient's Last Name]\n \
             [Select 3 to change Patient's Telephone Number]\n \
              [Select 4 to change Patient's Email Address]\n")
    
    select = str(input("Choice: "))
    
    if   select == "1":
        patient_info[1] = str(input("Enter New First Name: "))

        for data in sheet["A"]:
            for row in page.iter_rows(min_row=findpos(patient_info[0]), min_col=2, max_row=findpos(patient_info[0]), max_col=2):
                for cell in row:
                    cell.value = patient_info[1]
                    wb.save(filename=workbook)

    elif select == "2":
        for data in sheet["A"]:
            patient_info[2] = str(input("Enter New Last Name: "))
            for row in page.iter_rows(min_row=findpos(patient_info[0]), min_col=3, max_row=findpos(patient_info[0]), max_col=3):
                for cell in row:
                    cell.value = patient_info[2]
                    wb.save(filename=workbook)

    elif select == "3":
        patient_info[3] = str(input("Enter New Telephone Number: "))
        for data in sheet["A"]:
            for row in page.iter_rows(min_row=findpos(patient_info[0]), min_col=4, max_row=findpos(patient_info[0]), max_col=4):
                for cell in row:
                    cell.value = patient_info[3]
                    wb.save(filename=workbook)
                    
    elif select == "4":
        patient_info[4] = str(input("Enter New Email Address: "))
        for data in sheet["A"]:
            for row in page.iter_rows(min_row=findpos(patient_info[0]), min_col=5, max_row=findpos(patient_info[0]), max_col=5):
                for cell in row:
                    cell.value = patient_info[4]
                    wb.save(filename=workbook)

def delete_patient():
    choice = str(input("Choose patient ID to be deleted: "))
    page.delete_rows(findpos(choice))
    wb.save(filename=workbook)

def sort():
    df = pd.read_excel(workbook)
    print("Original Database")
    print(df)

    sorted = df.sort_values("ID")

    print ("Sorted")
    print(sorted)

                    
if __name__ == "__main__":
    #create_database("Patients_Database.xlsx")
    
    #add_patient()
    find_patient()
    #update_patient()
    #delete_patient()
    #sort()
